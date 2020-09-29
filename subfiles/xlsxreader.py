import pandas as pd
import numpy as np
from openpyxl import load_workbook


class Spreadsheet:
    def __init__(self, xlsx_path):
        """Initialising excel object which will read values through pandas library

        Args:
            xlsx_path (string): Path to xlsx file
        """
        self.path = xlsx_path
        self.xlsx_file_pd = pd.read_excel(self.path)
        self.xlsx_file_op = load_workbook(self.path)

    def get_xlsx_data(self):
        """Getting values from Excel spreadsheet

        Returns:
            [array]: Arrays of spreadsheet rows
        """
        return self.xlsx_file_pd

    def get_headers_list(self):
        return self.xlsx_file_pd.columns.values.tolist()

    def add_to_row(self, row_number, text):
        """Adding the value in the last column of the row in Excel spreadsheet. This function is intended to print result of email sending.

        Args:
            row_number (int): The number of the row in the Excel spreadsheet we want to insert value to. The first row has number 0 and is used for columns labales. Start with 1 for the first value.
            text (string): The value to insert in the cell in Excel spreadsheet.
        """
        workbook = self.xlsx_file_op.active
        workbook.cell(column=workbook.max_column,
                      row=row_number+1, value=text)
        self.xlsx_file_op.save(self.path)
